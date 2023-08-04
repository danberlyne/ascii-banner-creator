import openpyxl
from pathlib import Path

wb = openpyxl.load_workbook('entries.xlsx')
ws = wb.active
nations = [cell for cell in ws['A'] if cell.value != 'NATION']

for nation in nations:
    artist = ws[f'B{nation.row}']
    song = ws[f'C{nation.row}']
    file = open(Path('Flags-ASCII', f'{nation.value}.txt'))
    lines = [line.strip('\n') for line in file.readlines()]
    file.close()
    n = len(lines)
    if n % 2 == 0:
        lines[int(n/2 - 2)] += ' ' * 10 + f'{nation.value}'.upper()
        lines[int(n/2 - 1)] += ' ' * 10 + f'{artist.value}'
        lines[int(n/2)] += ' ' * 10 + f'{song.value}'
    else:
        lines[int((n-1)/2 - 1)] += ' ' * 10 + f'{nation.value}'.upper()
        lines[int((n+1)/2 - 1)] += ' ' * 10 + f'{artist.value}'
        lines[int((n+3)/2 - 1)] += ' ' * 10 + f'{song.value}'

    file = open(Path('Banners', f'{nation.value}.txt'), 'w')
    file.write('\n'.join(lines))
    file.close()