#!/usr/bin/env python3
# show_generator.py - Combines banners into a single text file in BBCode.

import openpyxl
from pathlib import Path

def main():
    wb = openpyxl.load_workbook('entries.xlsx')
    ws = wb.active
    nations = [cell for cell in ws['A'] if cell.value != 'NATION']

    show = []

    for nation in nations:
        url = ws[f'D{nation.row}']
        file = open(Path('Banners', f'{nation.value}.txt'))
        banner = ''.join(file.readlines())
        file.close()
        bbcode = f'[URL={url.value}][CODE]' + banner + '[/CODE][/URL]'
        show.append(bbcode)

    file = open('show.txt', 'w')
    file.write('\n'.join(show))
    file.close()

if __name__ == '__main__':
    main()